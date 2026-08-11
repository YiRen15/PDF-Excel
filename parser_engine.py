import os
import re
import shutil
import copy
import zipfile
import tempfile
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font
from concurrent.futures import ThreadPoolExecutor, as_completed

# 优先使用 C 语言编写的 10 倍极速 PyMuPDF 引擎，若未安装则自动回退至 pdfplumber
try:
    import fitz
    HAS_FITZ = True
except ImportError:
    HAS_FITZ = False
    import pdfplumber

def parse_hms(dur_str):
    if not dur_str:
        return 0, 0, 0
    dur_str = dur_str.strip()
    
    # 1. 格式 1: HH:MM:SS 或 MM:SS (例如 07:33:00 或 07:33 或 18:53)
    m_time = re.search(r'(\d{1,2})[:：](\d{1,2})(?:[:：](\d{1,2}))?', dur_str)
    if m_time:
        if m_time.group(3):
            return int(m_time.group(1)), int(m_time.group(2)), int(m_time.group(3))
        else:
            val1, val2 = int(m_time.group(1)), int(m_time.group(2))
            if val1 >= 24:
                return val1, val2, 0
            else:
                return 0, val1, val2

    # 2. 格式 2: 标准中英文单位 (含有 小时/时/h/H, 分钟/分/min/m/M, 秒/sec/s/S)
    hours, mins, secs = 0, 0, 0
    m_h = re.search(r'(\d+)\s*(?:小时|h|时)', dur_str, re.I)
    if m_h: hours = int(m_h.group(1))
    
    m_m = re.search(r'(\d+)\s*(?:分钟?|min|m|分)', dur_str, re.I)
    if m_m: mins = int(m_m.group(1))
    
    m_s = re.search(r'(\d+)\s*(?:秒|sec|s)', dur_str, re.I)
    if m_s: secs = int(m_s.group(1))
    
    if hours > 0 or mins > 0 or secs > 0:
        return hours, mins, secs

    # 3. 格式 3: 单双引号符号简写 (例如 18'53" 或 18′53″ 或 18' 或 53")
    m_sym = re.search(r"(?:(\d+)['′])?\s*(?:(\d+)[\"″])?", dur_str)
    if m_sym and (m_sym.group(1) or m_sym.group(2)):
        mins = int(m_sym.group(1)) if m_sym.group(1) else 0
        secs = int(m_sym.group(2)) if m_sym.group(2) else 0
        return 0, mins, secs

    # 4. 格式 4: 智能多数字保底推算 (例如 18 53 提取出 18 分 53 秒)
    nums = [int(n) for n in re.findall(r'\d+', dur_str)]
    if len(nums) >= 2:
        return 0, nums[0], nums[1]
    elif len(nums) == 1:
        return 0, 0, nums[0]
    
    return 0, 0, 0

def parse_single_pdf(pdf_path):
    """
    单文件 PDF 解析纯函数 (集成 PyMuPDF 8-10倍极速提取引擎)
    """
    try:
        filename = os.path.basename(pdf_path)
        text = ""
        annots = []
        
        if HAS_FITZ:
            doc = fitz.open(pdf_path)
            if len(doc) == 0:
                doc.close()
                return None
            page = doc[0]
            text = page.get_text("text", sort=True) or ""
            for annot in (page.annots() or []):
                info = annot.info
                if info and info.get("content"):
                    annots.append({"contents": info.get("content")})
            doc.close()
        else:
            with pdfplumber.open(pdf_path) as pdf:
                if len(pdf.pages) == 0:
                    return None
                page = pdf.pages[0]
                text = page.extract_text() or ""
                annots = getattr(page, "annots", []) or []

        text_no_space = re.sub(r'\s+', '', text)
        data = {}
        
        # 1. 受试者编号 (用户姓名) - 匹配到 '年龄'、'性别' 等字段为止
        m = re.search(r'(?:用户姓名|受试者姓名|患者姓名|姓名)[:：](.*?)(?=年龄|性别|病历号|床号|科室|报告日期)', text_no_space)
        if not m:
            m = re.search(r'(?:用户姓名|受试者姓名|患者姓名|姓名)[:：](.*)', text_no_space)
        data['SUBJID'] = m.group(1) if m and m.group(1) else ""
        
        # 2. 开始监测日期 (yyyy-MM-dd HH:mm)
        m = re.search(r'(?:记录日期|测试日期|监测日期|检测日期)[:：](\d{4}[-/.]\d{2}[-/.]\d{2})(\d{2}:\d{2})', text_no_space)
        if m:
            date_part = m.group(1).replace('/', '-').replace('.', '-')
            time_part = m.group(2)
            data['ECGSTDAT'] = f"{date_part} {time_part}"
        else:
            m = re.search(r'(?:记录日期|测试日期|监测日期|检测日期)[:：](\d{4}[-/.]\d{2}[-/.]\d{2})', text_no_space)
            data['ECGSTDAT'] = m.group(1).replace('/', '-').replace('.', '-') if m else ""
            
        # 3. 监测时长 (无数值时填 0，替代原斜杠 /)
        m = re.search(r'记录时间[:：](\d+)[:：](\d+)', text_no_space)
        if m:
            hours = int(m.group(1))
            mins = int(m.group(2))
            data['ECGDURH'] = f"{hours:02d}" if hours > 0 else 0
            data['ECGDURM'] = f"{mins:02d}" if mins > 0 else 0
        else:
            data['ECGDURH'] = 0
            data['ECGDURM'] = 0
            
        data['ECGDURH_U'] = "小时"
        data['ECGDURM_U'] = "分"
        
        # 5. 心率指标
        m_avg = re.search(r'平均心率[:：](\d+)[(（]?bpm[)）]?', text_no_space, re.I)
        data['ECGHR'] = int(m_avg.group(1)) if m_avg else "/"
        data['ECGHR_U'] = "bpm"
        
        m_max = re.search(r'(?:最大|最快)心率[:：](\d+)[(（]?bpm[)）]?', text_no_space, re.I)
        data['ECGHRMAX'] = int(m_max.group(1)) if m_max else "/"
        data['ECGHRMAX_U'] = "bpm"
        
        m_min = re.search(r'(?:最小|最慢)心率[:：](\d+)[(（]?bpm[)）]?', text_no_space, re.I)
        data['ECGHRMIN'] = int(m_min.group(1)) if m_min else "/"
        data['ECGHRMIN_U'] = "bpm"
        
        # 6. 最长RR间期
        m_rr = re.search(r'最大RR间期[:：](\d+)毫秒', text_no_space, re.I)
        if not m_rr:
            m_rr = re.search(r'最长R-R间期[:：]?(\d+)ms', text_no_space, re.I)
        data['ECGRR'] = int(m_rr.group(1)) if m_rr else "/"
        data['ECGRR_U'] = "ms"
        
        # 分析的总心搏数
        m_beats = re.search(r'分析的心搏数[:：](\d+)[(（]?次[)）]?', text_no_space)
        total_beats = int(m_beats.group(1)) if m_beats else 1
        
        # 7. 房颤基础数据提取 (切片隔离：绝不跨区扫描房扑/起搏)
        af_sec_m = re.search(r'房颤分析.*?(?=房扑分析|起搏分析|室上性节律|室性节律|ST段分析|结论|$)', text_no_space)
        af_sec = af_sec_m.group(0) if af_sec_m else text_no_space

        fl_sec_m = re.search(r'房扑分析.*?(?=结论|24小时数据图|24小时散点图|24小时趋势图|报告医生|起搏分析|$)', text_no_space)
        fl_sec = fl_sec_m.group(0) if fl_sec_m else text_no_space

        m_af_beats = re.search(r'房颤心搏[:：](\d+)[(（]?次[)）]?', af_sec)
        if not m_af_beats:
            m_af_beats = re.search(r'房颤心搏[:：](\d+)', af_sec)
        af_beats = int(m_af_beats.group(1)) if m_af_beats else 0
        
        af_dur_str = ""
        m_af_block = re.search(r'(?:持续时间|最长持续时间|时间)[:：]?(.*?)(?=发生次数|大于|$)', af_sec)
        if m_af_block:
            af_dur_str = m_af_block.group(1).strip()
            
        af_hours, af_mins, af_secs = parse_hms(af_dur_str)
        
        has_printed_afib_text = bool(re.search(r'(?:占总心搏|占比|负荷)[:：]?\s*(?:小于|＜|<|不足)\s*1(?:\.0)?', af_sec))
        m_af_pct_check = re.search(r'房颤心搏[:：]\d+[(（]?次[)）]?[,，]?\s*(?:占总心搏|占比|负荷)[:：]?(\d+(?:\.\d+)?)[(（]?[%％][)）]?', af_sec)
        if not m_af_pct_check:
            m_af_pct_check = re.search(r'(?:占比|占总心搏|总占比|负荷)[:：]?(\d+(?:\.\d+)?)[%％]', af_sec)
        has_printed_afib_val = bool(m_af_pct_check)
        
        # 8. 规则的房性心动过速 (室上速) 和 房扑
        svt_block_match = re.search(r'室上性节律.*?二联律:.*?室上速[:：](\d+)[(（]?阵[)）]?', text_no_space)
        if not svt_block_match:
            svt_block_match = re.search(r'室上速[:：](\d+)[(（]?阵[)）]?', text_no_space)
        svt_runs = int(svt_block_match.group(1)) if svt_block_match else 0
        
        m_svt_dur = re.search(r'最长持续时间(?:为)?(\d+)(?:s|秒)', text_no_space)
        svt_dur_sec = int(m_svt_dur.group(1)) if m_svt_dur else 0
        if svt_runs == 0:
            svt_dur_sec = 0
            
        svt_active = (svt_dur_sec >= 30)
        
        m_fl_beats = re.search(r'房扑心搏[:：](\d+)[(（]?次[)）]?,?\s*(?:占总心搏|占比|负荷)[:：]?(\d+(?:\.\d+)?)[(（]?[%％][)）]?', fl_sec)
        if m_fl_beats:
            fl_beats = int(m_fl_beats.group(1))
            fl_burden = float(m_fl_beats.group(2))
        else:
            m_fl_beats_only = re.search(r'房扑心搏[:：](\d+)', fl_sec)
            fl_beats = int(m_fl_beats_only.group(1)) if m_fl_beats_only else 0
            m_fl_pct_only = re.search(r'房扑(?:占比|占总心搏|总占比|负荷)[:：]?(\d+(?:\.\d+)?)[%％]', fl_sec)
            fl_burden = float(m_fl_pct_only.group(1)) if m_fl_pct_only else 0.0
            
        fl_dur_str = ""
        fl_block_match = re.search(r'(?:持续时间|最长持续时间|时间)[:：]?(.*?)(?=发生次数|大于|$)', fl_sec)
        if fl_block_match:
            fl_dur_str = fl_block_match.group(1).strip()
        
        fl_hours, fl_mins, fl_secs = parse_hms(fl_dur_str)
        fl_active = (fl_beats > 0 or fl_hours > 0 or fl_mins > 0)
        
        # 8.1 规则的房性心动过速 (室上速 + 房扑) 负荷精细化判定
        # 1. 房扑负荷原值/保底计算
        fl_burden_val = None
        if fl_active:
            if re.search(r'(?:占总心搏|占比|负荷)[:：]?\s*(?:小于|＜|<|不足)\s*1(?:\.0)?', fl_sec):
                fl_burden_val = "小于1"
            elif fl_burden > 0:
                if 0 < fl_burden < 1:
                    fl_burden_val = "小于1"
                else:
                    fl_burden_val = int(fl_burden) if fl_burden.is_integer() else fl_burden
            elif fl_beats > 0 and total_beats > 0:
                calc_fl_pct = (fl_beats / total_beats) * 100
                if calc_fl_pct < 1:
                    fl_burden_val = "小于1"
                else:
                    fl_burden_val = int(round(calc_fl_pct))

        # 2. 房速负荷原生打印值 (不自动公式计算)
        svt_burden_val = None
        if svt_active:
            for annot in annots:
                contents = annot.get("contents") if isinstance(annot, dict) else getattr(annot, "get", lambda k: None)("contents")
                if contents:
                    m_b = re.search(r'房速(?:总占比|占比)[:：]?(\d+(?:\.\d+)?)[%％]', str(contents))
                    if m_b:
                        svt_burden_val = float(m_b.group(1))
                        break
            if svt_burden_val is None:
                m_b2 = re.search(r'房速(?:总占比|占比)[:：]?(\d+(?:\.\d+)?)[%％]', text_no_space)
                if m_b2:
                    svt_burden_val = float(m_b2.group(1))

        # 9. 检查结果判定 (ECGORRES) - 第一步: 判定是否有房颤/规则房速/窦性心律
        conclusion_match = re.search(r'结论[:：]?(.*?)(?=24小时数据图|24小时散点图|24小时趋势图|报告医生|报告医师|$)', text_no_space)
        conclusion_text = conclusion_match.group(1) if conclusion_match else text_no_space
        
        # 9.1 从结论中精准切片出“房速/室上速专属语句”
        svt_clause_text = ""
        svt_clause_match = re.search(r'([^，,；;。\n]*?(?:房性心动过速|房速|室上速|室上性心动过速)[^。;\n]*?)(?=[。;\n]|$)', conclusion_text)
        if svt_clause_match:
            svt_clause_text = svt_clause_match.group(1)

        # 初始化异常告警标志与告警日志
        data['has_warning'] = False
        data['warnings'] = []

        # 9.2 仅在房速专属语句中提取最长持续时间
        concl_svt_dur_sec = 0
        if svt_clause_text:
            m_concl_dur = re.search(r'(?:最长持续时间|最长时间|最长一阵|持续时间)[：:为]?([0-9:a-zA-Z小时分秒钟]+)', svt_clause_text, re.I)
            if m_concl_dur:
                dur_raw = m_concl_dur.group(1)
                d_h, d_m, d_s = parse_hms(dur_raw)
                total_concl_sec = d_h * 3600 + d_m * 60 + d_s
                if total_concl_sec > 0:
                    concl_svt_dur_sec = total_concl_sec
                else:
                    m_num = re.search(r'(\d+)', dur_raw)
                    if m_num:
                        concl_svt_dur_sec = int(m_num.group(1))
                    if dur_raw:
                        data['has_warning'] = True
                        data['warnings'].append(f"发现未识别的持续时间格式 '{dur_raw}'，需人工核对")

        # 保底持续时间 (优先采用结论专属提取)
        effective_svt_dur_sec = max(concl_svt_dur_sec, svt_dur_sec)
        svt_valid = (effective_svt_dur_sec >= 30)

        # 8.2 持续时间相加与满60进位转化 (房速最长持续时间必须 >= 30秒 才允许参与换算与累加)
        svt_h = (effective_svt_dur_sec // 3600) if svt_valid else 0
        svt_m = ((effective_svt_dur_sec % 3600) // 60) if svt_valid else 0
        svt_s = (effective_svt_dur_sec % 60) if svt_valid else 0

        fl_h = fl_hours if fl_active else 0
        fl_m = fl_mins if fl_active else 0
        fl_s = fl_secs if fl_active else 0

        # 秒数进位
        total_s_raw = svt_s + fl_s
        s_carry_m = total_s_raw // 60
        at_final_s = total_s_raw % 60

        # 分钟进位
        total_m_raw = svt_m + fl_m + s_carry_m
        m_carry_h = total_m_raw // 60
        at_final_m = total_m_raw % 60

        # 小时数累加
        at_final_h = svt_h + fl_h + m_carry_h

        selected_codes = []
        
        # 规则 2: 房颤 (需排除 "未见房颤"、"无心房颤动"、"不排除房颤" 等否定句或警告提醒词)
        afib_negated = re.search(r'(未见|无|未发现|未检测到|不伴|否认|未出现|无明显|不排除|排除|待排|疑为|疑似)[^，,；;。\n]*?(房颤|心房颤动)', conclusion_text)
        has_afib_text = any(k in conclusion_text for k in ["心房颤动", "房颤"]) and not afib_negated
        has_afib_data = (af_beats > 0) or has_printed_afib_text or has_printed_afib_val
        if has_afib_text or has_afib_data:
            selected_codes.append("2")
            
        # 规则 3: 规则的房性心动过速 (房扑或室上速 >= 30s)
        has_aflutter_text = any(k in conclusion_text for k in ["心房扑动", "房扑"])
        has_aflutter_data = (fl_beats > 0) or (fl_burden > 0)
        cond3_1 = has_aflutter_text or has_aflutter_data
        
        svt_negated_in_concl = re.search(r'(未见|无|未发现|未检测到|不伴|否认|未出现|无明显|不排除|排除|待排|疑为|疑似)[^，,；;。\n]*?(房性心动过速|房速|室上速)', conclusion_text)
        has_svt_in_conclusion = any(k in conclusion_text for k in ["房性心动过速", "房速", "室上速"]) and not svt_negated_in_concl

        cond3_2 = (has_svt_in_conclusion or svt_active) and effective_svt_dur_sec >= 30
        if cond3_1 or cond3_2:
            selected_codes.append("3")
            
        # 规则 1: 窦性心律 (关键字精简包含: 窦性心律、窦性心动、起搏心律)
        sinus_keywords = ["窦性心律", "窦性心动", "起搏心律"]
        sinus_negated = re.search(r'(未见|无|未发现|未检测到|否认|无明显)[^，,；;。\n\d]*?(窦性心律|窦性心动|起搏心律)', conclusion_text)
        has_sinus = any(k in conclusion_text for k in sinus_keywords) and not sinus_negated
        has_any_af = ("2" in selected_codes) or ("3" in selected_codes)
        svt_under_30 = (effective_svt_dur_sec < 30)
        if has_sinus and not has_any_af and svt_under_30:
            selected_codes.append("1")
            
        # 需求三：中括号包裹与选择项文字映射
        if not selected_codes:
            code_raw = "/"
            code_bracket = "/"
            code_opt = "/"
        else:
            code_raw = ",".join(selected_codes)
            code_bracket = f"[{code_raw}]"
            
            if code_raw == "1":
                code_opt = "窦性心律，无心房颤动、规则的房性心动过速"
            elif code_raw == "2":
                code_opt = "房颤"
            elif code_raw == "3":
                code_opt = "规则的房性心动过速"
            elif "2" in code_raw and "3" in code_raw:
                code_opt = "房颤,规则的房性心动过速"
            else:
                code_opt = "/"

        data['ECGORRES'] = code_bracket
        data['ECGORRES_OPT'] = code_opt

        # 需求四：第二步——根据诊断判定结果，对应填充负荷与时长数据（非对应诊断列留空 ""）
        has_2 = "2" in code_raw
        has_3 = "3" in code_raw

        # 1. 房颤组字段处理
        if has_2:
            af_burden_val = None
            if has_printed_afib_text:
                af_burden_val = "小于1"
            elif m_af_pct_check:
                try:
                    raw_v = float(m_af_pct_check.group(1))
                    if 0 < raw_v < 1:
                        af_burden_val = "小于1"
                    else:
                        af_burden_val = int(raw_v) if raw_v.is_integer() else raw_v
                except ValueError:
                    af_burden_val = None
            
            if af_burden_val is None and af_beats > 0 and total_beats > 0:
                calc_af_pct = (af_beats / total_beats) * 100
                if calc_af_pct < 1:
                    af_burden_val = "小于1"
                else:
                    af_burden_val = int(round(calc_af_pct))

            if af_burden_val is None:
                af_burden_val = 0

            data['ECGAFBURD'] = af_burden_val
            data['ECGAFBURD_U'] = "%"
            data['ECGAFDURH'] = af_hours if af_hours > 0 else 0
            data['ECGAFDURH_U'] = "小时"
            data['ECGAFDUR'] = af_mins if af_mins > 0 else 0
            data['ECGAFDUR_U'] = "分"
        else:
            data['ECGAFBURD'] = ""
            data['ECGAFBURD_U'] = ""
            data['ECGAFDURH'] = ""
            data['ECGAFDURH_U'] = ""
            data['ECGAFDUR'] = ""
            data['ECGAFDUR_U'] = ""

        # 2. 规则心动过速组字段处理
        if has_3:
            # 触发房速人工检查提示的 3 个条件：1. 结论肯定有房速 2. 房速持续时间 >= 30s 3. 报告未打印房速占比原值
            needs_svt_manual_check = has_svt_in_conclusion and (effective_svt_dur_sec >= 30) and (svt_burden_val is None)

            if needs_svt_manual_check:
                data['has_warning'] = True
                data['warnings'].append("报告结论确诊有房速且持续时间 >= 30s，但报告未打印房速占比原值，需要人工核对")

            if fl_burden_val is not None and needs_svt_manual_check:
                fl_str = f"{int(fl_burden_val)}" if isinstance(fl_burden_val, (int, float)) and float(fl_burden_val).is_integer() else f"{fl_burden_val}"
                data['ECGATBURD'] = f"{fl_str}, 报告房速负荷需要人工检查"
                data['ECGATBURD_U'] = "%"
            elif fl_burden_val is None and needs_svt_manual_check:
                data['ECGATBURD'] = "报告房速负荷需要人工检查"
                data['ECGATBURD_U'] = "%"
            elif fl_burden_val is not None and svt_burden_val is not None:
                v_fl = 0 if fl_burden_val == "小于1" else float(fl_burden_val)
                total_v = v_fl + float(svt_burden_val)
                data['ECGATBURD'] = int(total_v) if float(total_v).is_integer() else round(total_v, 2)
                data['ECGATBURD_U'] = "%"
            elif svt_burden_val is not None:
                data['ECGATBURD'] = int(svt_burden_val) if float(svt_burden_val).is_integer() else round(svt_burden_val, 2)
                data['ECGATBURD_U'] = "%"
            elif fl_burden_val is not None:
                data['ECGATBURD'] = int(fl_burden_val) if isinstance(fl_burden_val, (int, float)) and float(fl_burden_val).is_integer() else fl_burden_val
                data['ECGATBURD_U'] = "%"
            else:
                data['ECGATBURD'] = "报告房速负荷需要人工检查"
                data['ECGATBURD_U'] = "%"

            data['ECGATDURH'] = at_final_h if at_final_h > 0 else 0
            data['ECGATDURH_U'] = "小时"
            data['ECGATDUR'] = at_final_m if at_final_m > 0 else 0
            data['ECGATDUR_U'] = "分"
            data['ECGATDURS'] = at_final_s if at_final_s > 0 else 0
            data['ECGATDURS_U'] = "秒"
        else:
            data['ECGATBURD'] = ""
            data['ECGATBURD_U'] = ""
            data['ECGATDURH'] = ""
            data['ECGATDURH_U'] = ""
            data['ECGATDUR'] = ""
            data['ECGATDUR_U'] = ""
            data['ECGATDURS'] = ""
            data['ECGATDURS_U'] = ""

        data['ECGAFR'] = "/"
        data['ECGAFR_OPT'] = "/"
        
        data['_filename'] = filename
        return data

    except Exception as e:
        print(f"解析文件 {pdf_path} 出错: {e}")
        return {
            'SUBJID': os.path.splitext(os.path.basename(pdf_path))[0],
            'ECGORRES': "解析失败",
            '_filename': os.path.basename(pdf_path),
            '_error': str(e)
        }

def parse_pdf_batch(pdf_paths, max_workers=32):
    total = len(pdf_paths)
    if total == 0:
        return []
    
    results = []
    
    if total == 1:
        res = parse_single_pdf(pdf_paths[0])
        if res: results.append(res)
        return results

    workers = min(max_workers, min(32, max(4, total)))
    with ThreadPoolExecutor(max_workers=workers) as executor:
        future_to_path = {executor.submit(parse_single_pdf, path): path for path in pdf_paths}
        for future in as_completed(future_to_path):
            try:
                res = future.result()
                if res:
                    results.append(res)
            except Exception as exc:
                path = future_to_path[future]
                print(f"{path} 处理时触发异常: {exc}")
                
    results = sorted(results, key=lambda x: str(x.get('SUBJID') or x.get('_filename') or ''))
    return results

def write_all_to_excel(data_list, template_path, output_path):
    shutil.copy(template_path, output_path)
    
    wb = openpyxl.load_workbook(output_path)
    sheet = wb['表头整理建议'] if '表头整理建议' in wb.sheetnames else wb.active
    
    try:
        sheet.unmerge_cells('K2:L2')
    except Exception:
        pass
        
    col_mapping = {
        "受试者编号": "SUBJID",
        "开始监测日期": "ECGSTDAT",
        "监测时长": "ECGDURH",
        "监测时长_单位": "ECGDURH_U",
        "监测时长（分）": "ECGDURM",
        "监测时长（分）_单位": "ECGDURM_U",
        "监测时长（分）_单": "ECGDURM_U",
        "分析时长": "ECGANADURH",
        "分析时长_单位": "ECGANADURH_U",
        "分析时长（分）": "ECGANADURM",
        "分析时长（分）_单位": "ECGANADURM_U",
        "检查结果": "ECGORRES",
        "检查结果_选择项": "ECGORRES_OPT",
        "平均心率": "ECGHR",
        "平均心率_单位": "ECGHR_U",
        "最大心率": "ECGHRMAX",
        "最大心率_单位": "ECGHRMAX_U",
        "最小心率": "ECGHRMIN",
        "最小心率_单位": "ECGHRMIN_U",
        "最长RR间期": "ECGRR",
        "最长RR间期_单位": "ECGRR_U",
        "房颤负荷": "ECGAFBURD",
        "房颤负荷_单位": "ECGAFBURD_U",
        "房颤负荷最长持续时间（小时）": "ECGAFDURH",
        "房颤负荷最长持续时间（小时）_单位": "ECGAFDURH_U",
        "房颤负荷最长持续时间": "ECGAFDUR",
        "房颤负荷最长持续时间_单位": "ECGAFDUR_U",
        "规则的房性心动过速负荷": "ECGATBURD",
        "规则的房性心动过速负荷_单位": "ECGATBURD_U",
        "规则的房性心动过速负荷最长持续时间（小时）": "ECGATDURH",
        "规则的房性心动过速负荷最长持续时间（小时）_单位": "ECGATDURH_U",
        "规则的房性心动过速负荷最长持续时间": "ECGATDUR",
        "规则的房性心动过速负荷最长持续时间_单位": "ECGATDUR_U",
        "规则的房性心动过速负荷最长持续时间（秒）": "ECGATDURS",
        "规则的房性心动过速负荷最长持续时间（秒）_单位": "ECGATDURS_U"
    }

    # 智能识别表头行数与数据起写行 (默认从第 3 行写入，直接覆盖第 3 行样例行)
    r1_val = str(sheet.cell(row=1, column=1).value or '').strip()
    r2_val = str(sheet.cell(row=2, column=1).value or '').strip()
    r3_val = str(sheet.cell(row=3, column=1).value or '').strip()

    start_data_row = 3
    if r1_val == 'SUBJID' and r2_val == '受试者编号' and r3_val == '姓名':
        start_data_row = 3
    elif r1_val == '受试者编号':
        start_data_row = 2

    # 解析各列字段 Key
    max_cols = sheet.max_column
    col_keys = {}
    for col_idx in range(1, max_cols + 1):
        c1 = str(sheet.cell(row=1, column=col_idx).value or '').strip()
        c2 = str(sheet.cell(row=2, column=col_idx).value or '').strip()
        c3 = str(sheet.cell(row=3, column=col_idx).value or '').strip()

        key = None
        if c1 in col_mapping.values():
            key = c1
        elif c2 in col_mapping:
            key = col_mapping[c2]
        elif c1 in col_mapping:
            key = col_mapping[c1]
        elif c3 in col_mapping:
            key = col_mapping[c3]
            
        if key:
            col_keys[col_idx] = key

    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    black_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    for idx, data_dict in enumerate(data_list):
        row_idx = start_data_row + idx
        for col_idx, key in col_keys.items():
            cell = sheet.cell(row=row_idx, column=col_idx)
            val = data_dict.get(key, "")
            cell.value = val
            if key == 'ECGORRES_OPT':
                cell.alignment = left_alignment
            else:
                cell.alignment = center_alignment
            cell.font = Font(name='等线', size=11, bold=False, italic=False, color='000000')
            cell.border = black_border

    wb.save(output_path)
    wb.close()
    return output_path

def extract_zip(zip_path, extract_to):
    pdf_paths = []
    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
        zip_ref.extractall(extract_to)
        
    for root, _, files in os.walk(extract_to):
        for f in files:
            if f.lower().endswith('.pdf') and not f.startswith('._'):
                pdf_paths.append(os.path.join(root, f))
    return pdf_paths

import os
import re
import shutil
import glob
import copy
import pdfplumber
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font

def parse_hms(dur_str):
    hours = 0
    mins = 0
    secs = 0
    m = re.search(r'(\d+)小时', dur_str)
    if m: hours = int(m.group(1))
    m = re.search(r'(\d+)分钟', dur_str)
    if m: mins = int(m.group(1))
    m = re.search(r'(\d+)秒', dur_str)
    if m: secs = int(m.group(1))
    return hours, mins, secs

def parse_pdf(pdf_path):
    print(f"解析 PDF 文件: {os.path.basename(pdf_path)}")
    with pdfplumber.open(pdf_path) as pdf:
        page = pdf.pages[0]
        text = page.extract_text()
        annots = getattr(page, "annots", [])
    
    # 移除所有空白字符（空格、换行、制表符等）以防换行或格式问题影响中文正则匹配
    text_no_space = re.sub(r'\s+', '', text)
    
    data = {}
    
    # 1. 受试者编号 (用户姓名) - 匹配到 '年龄' 为止
    m = re.search(r'用户姓名:(.*?)(?=年龄)', text_no_space)
    data['SUBJID'] = m.group(1) if m else ""
    
    # 2. 开始监测日期 (开始监测日期: yyyy-MM-dd HH:mm)
    m = re.search(r'记录日期:(\d{4}/\d{2}/\d{2})(\d{2}:\d{2})', text_no_space)
    if m:
        date_part = m.group(1).replace('/', '-')
        time_part = m.group(2)
        data['ECGSTDAT'] = f"{date_part} {time_part}"
    else:
        m = re.search(r'记录日期:(\d{4}/\d{2}/\d{2})', text_no_space)
        data['ECGSTDAT'] = m.group(1).replace('/', '-') if m else ""
        
    # 3. 监测时长 (记录时间)
    m = re.search(r'记录时间:(\d+):(\d+)', text_no_space)
    if m:
        hours = int(m.group(1))
        mins = int(m.group(2))
        data['ECGDURH'] = f"{hours:02d}" if hours > 0 else "/"
        data['ECGDURM'] = f"{mins:02d}" if mins > 0 else "/"
    else:
        data['ECGDURH'] = "/"
        data['ECGDURM'] = "/"
        
    data['ECGDURH_U'] = "小时"
    data['ECGDURM_U'] = "分"
    
    # 4. 分析时长 (分析时间)
    m = re.search(r'分析时间:(\d+):(\d+)', text_no_space)
    if m:
        hours = int(m.group(1))
        mins = int(m.group(2))
        data['ECGANADURH'] = f"{hours:02d}" if hours > 0 else "/"
        data['ECGANADURM'] = f"{mins:02d}" if mins > 0 else "/"
    else:
        data['ECGANADURH'] = "/"
        data['ECGANADURM'] = "/"
        
    data['ECGANADURH_U'] = "小时"
    data['ECGANADURM_U'] = "分"
    
    # 5. 心率指标
    m_avg = re.search(r'平均心率:(\d+)\(bpm\)', text_no_space)
    data['ECGHR'] = int(m_avg.group(1)) if m_avg else "/"
    data['ECGHR_U'] = "bpm"
    
    m_max = re.search(r'最大心率:(\d+)\(bpm\)', text_no_space)
    if not m_max:
        m_max = re.search(r'最快心率:(\d+)\(bpm\)', text_no_space)
    data['ECGHRMAX'] = int(m_max.group(1)) if m_max else "/"
    data['ECGHRMAX_U'] = "bpm"
    
    m_min = re.search(r'最小心率:(\d+)\(bpm\)', text_no_space)
    if not m_min:
        m_min = re.search(r'最慢心率:(\d+)\(bpm\)', text_no_space)
    data['ECGHRMIN'] = int(m_min.group(1)) if m_min else "/"
    data['ECGHRMIN_U'] = "bpm"
    
    # 6. 最长RR间期
    m_rr = re.search(r'最大RR间期:(\d+)毫秒', text_no_space)
    if not m_rr:
        m_rr = re.search(r'最长R-R间期(\d+)ms', text_no_space)
    data['ECGRR'] = int(m_rr.group(1)) if m_rr else "/"
    data['ECGRR_U'] = "ms"
    
    # 分析的总心搏数
    m_beats = re.search(r'分析的心搏数:(\d+)\(次\)', text_no_space)
    total_beats = int(m_beats.group(1)) if m_beats else 1
    
    # 7. 房颤分析
    m_af_beats = re.search(r'房颤心搏:(\d+)', text_no_space)
    af_beats = int(m_af_beats.group(1)) if m_af_beats else 0
    
    m_af_dur = re.search(r'房颤分析.*?持续时间:(\d+)', text_no_space)
    af_dur_sec = int(m_af_dur.group(1)) if m_af_dur else 0
    
    if af_beats > 0:
        data['ECGAFBURD'] = round((af_beats / total_beats) * 100, 2)
    else:
        data['ECGAFBURD'] = "/"
    data['ECGAFBURD_U'] = "%"
    
    if af_dur_sec > 0:
        af_hours = af_dur_sec // 3600
        af_mins = (af_dur_sec % 3600) // 60
        data['ECGAFDURH'] = f"{af_hours:02d}" if af_hours > 0 else "/"
        data['ECGAFDUR'] = f"{af_mins:02d}" if af_mins > 0 else "/"
    else:
        data['ECGAFDURH'] = "/"
        data['ECGAFDUR'] = "/"
        
    data['ECGAFDURH_U'] = "小时"
    data['ECGAFDUR_U'] = "分"
    
    # 8. 规则的房性心动过速 (室上速) 和 房扑
    svt_block_match = re.search(r'室上性节律.*?二联律:.*?室上速:(\d+)\(阵\)', text_no_space)
    if not svt_block_match:
        svt_block_match = re.search(r'室上速:(\d+)\(阵\)', text_no_space)
    svt_runs = int(svt_block_match.group(1)) if svt_block_match else 0
    
    m_svt_dur = re.search(r'最长持续时间(?:为)?(\d+)(?:s|秒)', text_no_space)
    svt_dur_sec = int(m_svt_dur.group(1)) if m_svt_dur else 0
    if svt_runs == 0:
        svt_dur_sec = 0
        
    # 房速持续时间是否 >= 30s
    svt_active = (svt_dur_sec >= 30)
    
    # 解析房扑
    m_fl_beats = re.search(r'房扑心搏:(\d+)\(次\),占总心搏(\d+)\(%\)', text_no_space)
    if m_fl_beats:
        fl_beats = int(m_fl_beats.group(1))
        fl_burden = int(m_fl_beats.group(2))
    else:
        m_fl_beats_only = re.search(r'房扑心搏:(\d+)', text_no_space)
        fl_beats = int(m_fl_beats_only.group(1)) if m_fl_beats_only else 0
        fl_burden = 0
        
    fl_dur_str = ""
    fl_block_match = re.search(r'房扑分析.*?持续时间:(.*?)发生次数', text_no_space)
    if fl_block_match:
        fl_dur_str = fl_block_match.group(1).strip()
    
    fl_hours, fl_mins, fl_secs = parse_hms(fl_dur_str)
    fl_active = (fl_beats > 0 or fl_hours > 0 or fl_mins > 0)
    
    # 8.1 规则的房性心动过速负荷
    svt_burden_val = "/"
    if svt_active:
        # 从 PDF 批注中提取“房速总占比XX%”
        for annot in annots:
            contents = annot.get("contents")
            if contents:
                m_b = re.search(r'房速总占比(\d+(?:\.\d+)?)%', contents)
                if m_b:
                    svt_burden_val = f"{m_b.group(1)}%"
                    break
        # 若未找到批注，自动计算
        if svt_burden_val == "/":
            m_svt_total = re.search(r'室上性节律.*?总数:(\d+)\(次\)', text_no_space)
            svt_total = int(m_svt_total.group(1)) if m_svt_total else 0
            m_svt_single = re.search(r'室上性节律.*?单发:(\d+)\(次\)', text_no_space)
            svt_single = int(m_svt_single.group(1)) if m_svt_single else 0
            m_svt_pair = re.search(r'室上性节律.*?成对:(\d+)\(阵\)', text_no_space)
            svt_pair = int(m_svt_pair.group(1)) if m_svt_pair else 0
            at_beats = max(0, svt_total - svt_single - svt_pair * 2)
            pct = round((at_beats / total_beats) * 100, 2)
            svt_burden_val = f"{pct}%" if pct > 0 else "/"
            
    fl_burden_val = f"{fl_burden}%" if fl_burden > 0 else "/"
    if not fl_active:
        fl_burden_val = "/"
        
    if svt_burden_val == "/" and fl_burden_val == "/":
        data['ECGATBURD'] = "/"
    else:
        data['ECGATBURD'] = f"1、 {svt_burden_val}\n2、 {fl_burden_val}"
    data['ECGATBURD_U'] = "%"
    
    # 8.2 SVT & Flutter 最长持续时间 (时/分/秒)
    # 小时
    svt_h_str = f"{svt_dur_sec // 3600:02d}" if (svt_active and (svt_dur_sec // 3600) > 0) else "/"
    fl_h_str = f"{fl_hours:02d}" if (fl_active and fl_hours > 0) else "/"
    if svt_h_str == "/" and fl_h_str == "/":
        data['ECGATDURH'] = "/"
    else:
        data['ECGATDURH'] = f"1、 {svt_h_str}\n2、 {fl_h_str}"
    data['ECGATDURH_U'] = "小时"
    
    # 分钟
    svt_m_str = f"{(svt_dur_sec % 3600) // 60:02d}" if (svt_active and ((svt_dur_sec % 3600) // 60) > 0) else "/"
    fl_m_str = f"{fl_mins:02d}" if (fl_active and fl_mins > 0) else "/"
    if svt_m_str == "/" and fl_m_str == "/":
        data['ECGATDUR'] = "/"
    else:
        data['ECGATDUR'] = f"1、 {svt_m_str}\n2、 {fl_m_str}"
    data['ECGATDUR_U'] = "分"
    
    # 秒
    svt_s_str = f"{svt_dur_sec % 60:02d}" if (svt_active and (svt_dur_sec % 60) > 0) else "/"
    fl_s_str = f"{fl_secs:02d}" if (fl_active and fl_secs > 0) else "/"
    if svt_s_str == "/" and fl_s_str == "/":
        data['ECGATDURS'] = "/"
    else:
        data['ECGATDURS'] = f"1、 {svt_s_str}\n2、 {fl_s_str}"
    data['ECGATDURS_U'] = "秒"
    
    # 9. 检查结果判定 (ECGORRES)
    conclusion_match = re.search(r'结论(.*?)报告医生', text_no_space)
    conclusion_text = conclusion_match.group(1) if conclusion_match else ""
    
    selected_codes = []
    
    # 规则 2: 房颤
    # 使用正则表达式以兼容“阵发性心房颤动”和“持续性心房颤动”等缩写
    has_afib = re.search(r'(阵发性|持续性|持续).*?(房颤|心房颤动)', conclusion_text)
    if "异位心律" in conclusion_text and has_afib:
        selected_codes.append("2")
        
    # 规则 3: 规则的房性心动过速
    # 房扑条件：异位心律且包含 阵发性/持续性 房扑/心房扑动
    has_aflutter = re.search(r'(阵发性|持续性|持续).*?(房扑|心房扑动)', conclusion_text)
    cond3_1 = "异位心律" in conclusion_text and has_aflutter
    cond3_2 = ("房性心动过速" in conclusion_text or "房速" in conclusion_text) and svt_dur_sec >= 30
    if cond3_1 or cond3_2:
        selected_codes.append("3")
        
    # 规则 1: 窦性心律
    has_sinus = "窦性心律" in conclusion_text
    has_af = any(x in conclusion_text for x in ["心房颤动", "房颤", "心房扑动", "房扑"])
    svt_under_30 = (svt_dur_sec < 30)
    if has_sinus and not has_af and svt_under_30:
        selected_codes.append("1")
        
    if not selected_codes:
        data['ECGORRES'] = "/"
    else:
        data['ECGORRES'] = ",".join(selected_codes)
        
    # 根据用户要求，检查结果_选择项固定保留显示所有的123选项选项（即不改动原来的多行定义文本）
    data['ECGORRES_OPT'] = "1=窦性心律，无心房颤动、规则的房性心动过速;\n2=房颤;\n3=规则的房性心动过速"
        
    data['ECGAFR'] = "/"
    data['ECGAFR_OPT'] = "/"
    
    return data

def write_all_to_excel(data_list, template_path, output_path):
    print(f"备份模板并生成合并的 Excel 表格: {os.path.basename(output_path)}")
    shutil.copy(template_path, output_path)
    
    wb = openpyxl.load_workbook(output_path)
    sheet = wb['表头整理建议']
    
    # 检查并取消合并第二行的 K2:L2 (检查结果 与 检查结果_选择项)
    try:
        sheet.unmerge_cells('K2:L2')
        print("已取消 K2:L2 单元格合并，以便独立填写检查结果两列。")
    except ValueError:
        pass
        
    col_mapping = {
        "受试者编号": "SUBJID",
        "开始监测日期": "ECGSTDAT",
        "监测时长": "ECGDURH",
        "监测时长_单位": "ECGDURH_U",
        "监测时长（分）": "ECGDURM",
        "监测时长（分）_单位": "ECGDURM_U",
        "分析时长": "ECGANADURH",
        "分析时长_单位": "ECGANADURH_U",
        "分析时长（分）": "ECGANADURM",
        "分析时长（分）_单位": "ECGANADURM_U",
        "检查结果": "ECGORRES",
        "检查结果_选择项": "ECGORRES_OPT",
        "平均心率": "ECGHR",
        "平均心率_单位": "ECGHR_U",
        "最大心率": "ECGHRMAX",
        "最大心率_单位": "ECGHRMAX_U",
        "最小心率": "ECGHRMIN",
        "最小心率_单位": "ECGHRMIN_U",
        "最长RR间期": "ECGRR",
        "最长RR间期_单位": "ECGRR_U",
        "房颤负荷": "ECGAFBURD",
        "房颤负荷_单位": "ECGAFBURD_U",
        "房颤负荷最长持续时间（小时）": "ECGAFDURH",
        "房颤负荷最长持续时间（小时）_单位": "ECGAFDURH_U",
        "房颤负荷最长持续时间": "ECGAFDUR",
        "房颤负荷最长持续时间_单位": "ECGAFDUR_U",
        "规则的房性心动过速负荷": "ECGATBURD",
        "规则的房性心动过速负荷_单位": "ECGATBURD_U",
        "规则的房性心动过速负荷最长持续时间（小时）": "ECGATDURH",
        "规则的房性心动过速负荷最长持续时间（小时）_单位": "ECGATDURH_U",
        "规则的房性心动过速负荷最长持续时间": "ECGATDUR",
        "规则的房性心动过速负荷最长持续时间_单位": "ECGATDUR_U",
        "规则的房性心动过速负荷最长持续时间（秒）": "ECGATDURS",
        "规则的房性心动过速负荷最长持续时间（秒）_单位": "ECGATDURS_U"
    }
    
    headers = [cell.value for cell in sheet[1]]
    
    # 缓存第二行的原始单元格样式，以便后续各行复用
    row2_styles = {}
    for col_idx in range(1, len(headers) + 1):
        cell = sheet.cell(row=2, column=col_idx)
        row2_styles[col_idx] = {
            'font': copy.copy(cell.font) if cell.font else None,
            'fill': copy.copy(cell.fill) if cell.fill else None,
            'border': copy.copy(cell.border) if cell.border else None,
            'number_format': cell.number_format,
            'protection': copy.copy(cell.protection) if cell.protection else None
        }
    
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # 定义全实线黑框
    black_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # 填充每一行数据，从第二行开始写入
    for idx, data_dict in enumerate(data_list):
        row_idx = 2 + idx
        for col_idx, h1 in enumerate(headers, 1):
            if not h1:
                continue
            h1 = h1.strip()
            
            # 创建/获取当前单元格
            cell = sheet.cell(row=row_idx, column=col_idx)
            
            # 1. 写入数值
            key = col_mapping.get(h1)
            val = data_dict.get(key) if key else None
            cell.value = val
            
            # 2. 复用第二行的样式格式，同时确保字体颜色强制为黑色 (RGB 000000)
            style = row2_styles.get(col_idx)
            if style:
                if style['font']:
                    # 复制字体样式，但将颜色修改为黑色
                    cell.font = Font(
                        name=style['font'].name,
                        size=style['font'].size,
                        bold=style['font'].bold,
                        italic=style['font'].italic,
                        charset=style['font'].charset,
                        family=style['font'].family,
                        underline=style['font'].underline,
                        strike=style['font'].strike,
                        color='000000'
                    )
                else:
                    cell.font = Font(color='000000')
                    
                if style['fill']: cell.fill = copy.copy(style['fill'])
                if style['number_format']: cell.number_format = style['number_format']
                if style['protection']: cell.protection = copy.copy(style['protection'])
            else:
                cell.font = Font(color='000000')
                
            # 3. 设置边框为所有黑线细框 (覆盖可能存在的无边框或浅色边框样式)
            cell.border = black_border
            
            # 4. 设置对齐方式（除了“检查结果_选择项”靠左对齐外，其他全部居中对齐）
            if h1 == '检查结果_选择项':
                cell.alignment = left_alignment
            else:
                cell.alignment = center_alignment
                
    wb.save(output_path)
    wb.close()
    print("写入完成。")

def main():
    # 获取脚本当前所在的目录，确保路径自适应，不管谁在其 Mac 上运行均可无缝定位
    script_dir = os.path.dirname(os.path.abspath(__file__))
    
    # 优先使用“随访心电记录仪检测记录_空模板.xlsx”防止模板数据被写入污染
    template_name = "随访心电记录仪检测记录_空模板.xlsx"
    template_path = os.path.join(script_dir, template_name)
    
    if not os.path.exists(template_path):
        # 兼容旧的文件命名
        template_path = os.path.join(script_dir, "随访心电记录仪检测记录.xlsx")
        if not os.path.exists(template_path):
            print(f"未找到模板文件，请确保 '随访心电记录仪检测记录_空模板.xlsx' 存在于以下目录：\n{script_dir}")
            return
        
    pdf_files = glob.glob(os.path.join(script_dir, "*.pdf"))
    if not pdf_files:
        print(f"未在当前目录下找到任何 PDF 报告。")
        return
        
    print(f"共找到 {len(pdf_files)} 个 PDF 文件进行合并解析...")
    pdf_files = sorted(pdf_files)
    
    data_list = []
    for pdf_path in pdf_files:
        try:
            data = parse_pdf(pdf_path)
            data_list.append(data)
        except Exception as e:
            print(f"解析 {pdf_path} 时出错: {e}")
            
    # 输出的文件名固定为“随访心电记录仪检测记录.xlsx”
    output_path = os.path.join(script_dir, "随访心电记录仪检测记录.xlsx")
    
    write_all_to_excel(data_list, template_path, output_path)

if __name__ == "__main__":
    main()
